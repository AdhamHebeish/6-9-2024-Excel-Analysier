from flask import Flask, render_template, request
import openpyxl
import io
from werkzeug.datastructures import ImmutableDict

app = Flask(__name__)
wsgi_app = app.wsgi_app

@app.route('/', methods=['POST', 'GET'])
def index():
    return render_template('index.html')

@app.route('/analyse', methods=['GET', 'POST'])
def analyse():
    fname = ""
    full = ""
    bd = ""
    num = ""
    err = ""
    file = request.files['file']
    if not file.filename.lower().endswith(('.xls', '.xlsx')): 
        err = 'You have not uploaded an excel file'
    else:
        file_data = io.BytesIO(file.read())
        file_data.seek(0)
        workbook = openpyxl.load_workbook(file_data)
        worksheet = workbook.active
        header_row = [cell.value.strip() for cell in next(worksheet.iter_rows(min_row=1, max_row=1))] 
        for row in worksheet.iter_rows(min_row=2):
            fname = row[header_row.index('First Name')].value if row[header_row.index('First Name')].value != None else ""
            full = row[header_row.index('Full Name')].value if row[header_row.index('Full Name')].value != None else ""
            bd = row[header_row.index('Birth Date')].value if row[header_row.index('Birth Date')].value != None else ""
            num = row[header_row.index('Mobile Number')].value if row[header_row.index('Mobile Number')].value != None else ""
        
    return render_template('analyse.html', err=err, full=full, bd=bd, fname=fname, num=num)

if __name__ == '__main__':
    app.run()